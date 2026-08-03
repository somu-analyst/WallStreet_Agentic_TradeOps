# -*- coding: utf-8 -*-
"""Build docs/IDEA_TRACKER.xlsx — one row per idea/question ever raised.

Re-runnable: edit ROWS below (or append) and re-run. Existing user edits in the sheet are
NOT preserved, so treat this script as the source of truth and add new items here.

    python tools/build_idea_tracker.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "docs", "IDEA_TRACKER.xlsx")

COLS = ["ID", "Date Added", "Raised By", "Category", "Idea / Question",
        "Detail", "Status", "Date Actioned", "Outcome / Evidence", "Commit",
        "Priority", "Next Step"]

# Status vocabulary: DONE | QUEUED | BLOCKED | REJECTED | WITHDRAWN | USER-ACTION
ROWS = [
# ---------------------------------------------------------------- security
(1,"2026-07-31","User","Security","Telegram bot token leaked on GitHub",
 "Secret scanning flagged 5 commits; live token 8407478799 exposed since 2025-12-23",
 "DONE","2026-07-31","Revoked via BotFather; new token installed; bot restarted, 0 auth errors","-","P0","-"),
(2,"2026-07-31","Claude","Security","Second bot @CE448_bot (8018716820) also leaked",
 "Token in 3 public commits; verified STILL ALIVE via getMe",
 "CLOSED","2026-08-02",
 "NOT THE USER'S BOT - closed on their instruction. Token still answers getMe (verified 2026-08-02), but it belongs to someone else so there is no revoke path from here. Its presence in this repo history remains a dangling-blob artifact only","-","P0",
 "You: @BotFather -> /revoke or /deletebot -> @CE448_bot"),
(3,"2026-07-31","User","Security","Purge git history / dangling commit 24e2ab8",
 "Asked whether to rewrite history to remove leaked tokens",
 "REJECTED","2026-07-31","filter-repo CANNOT reach dangling objects (not in clone); only GitHub Support can. Token now revoked = dead string. Rewrite would break 592 SHAs for cosmetics","-","P3",
 "Optional: dismiss GitHub alerts as 'revoked'"),
(4,"2026-07-30","Claude","Security","Full-history credential scan (all secret types)",
 "Scan for AWS/Anthropic/OpenAI/GitHub/Slack keys across 592 commits",
 "DONE","2026-08-02",
 "COMPLETE. Walked OBJECTS once (2,974 blobs <300KB) instead of walking commits - that is why the two earlier attempts timed out. Patterns: Telegram, AWS AKIA, sk-ant, OpenAI sk-, GitHub gh*_, Slack xox*, Google AIza. Result: ZERO AWS/Anthropic/OpenAI/GitHub/Slack/Google keys anywhere in history. Only 2 distinct Telegram tokens, both in UNREACHABLE (dangling) blobs: the pre-revoke token for bot 8407478799 (confirmed dead, 401) and @CE448_bot 8018716820 (confirmed still LIVE). The token currently in token.txt is live (@Analyst_Somu_AI_bot) and does NOT appear anywhere in git history","-","P2",
 "Re-run bounded to small files, in background"),
# ---------------------------------------------------------------- methodology
(5,"2026-07-31","Claude","Methodology","Pooled rank-IC t-stat is invalid on this data",
 "Random alpha expressions tested against the harness",
 "DONE","2026-07-31","68-79% of RANDOM alphas passed p<0.05 (best |t|=10.80); correct daily-IC test passes 0%. Cause: overlap + cross-correlation inflate t ~10x","5d0fa27","P0","-"),
(6,"2026-07-30","Claude","Strategy","/debate Technical agent weight 1.0 -> 1.3",
 "Raised on a backtest reporting rank-IC +0.047, t=+7.05",
 "WITHDRAWN","2026-07-31","REVERTED. That t-stat came from the broken pooled method. Correct test: t=+1.04 (5d) / +1.10 (10d) = no edge","5d0fa27","P0","-"),
(7,"2026-07-31","Claude","Methodology","Re-test 3 caveated significance claims",
 "Market Radar 'QQQ corr +0.37', Rotation 'IC +0.14', positioning 'IC +0.03' all used pooled stat",
 "DONE","2026-08-02","Re-tested under daily-IC. Rotation 'IC +0.14 p<1e-7' WITHDRAWN (t=-0.28 5d / +0.41 10d, 33 ETFs x 1500d). Market Radar '+0.37 p<0.001' DOWNGRADED to suggestive (104 overlapping days = ~21 independent; p=0.10; sign flips +0.36..-0.17 across sampling offsets). Skew panel UNTESTABLE (21 dates -> 4 independent). Both UI captions corrected and Playwright-verified","5d0fa27","P1",
 "Re-run each with daily cross-sectional IC"),
(8,"2026-07-31","Claude","Methodology","Walk-forward convention (enforced train/test)",
 "Backtests are done ad-hoc; no enforced standard",
 "DONE","2026-08-02",
 "Built tools/walkforward.py as the standard: expanding-window folds, parameter REFIT inside each fold on train only, train window stopped `horizon` short so its labels cannot peek into test, test ICs de-overlapped before any t-test, judgement on pooled TEST folds only. Convention written into .claude/rules/bot-conventions.md. Self-test on 120 tickers x 1400 days: momentum with a per-fold refit scores train IC +0.0179 vs test +0.0113, pooled t=+0.75 p=0.49 - does NOT survive; fold 5 alone reads +0.066 (t=+1.89), which is what an ad-hoc backtest would have reported. Sanity gate 2.5% (expect 5%)","-","P1","-"),
# ---------------------------------------------------------------- rec engine
(9,"2026-07-31","User","Strategy","Backtest all settled recommendations",
 "311 settled recs (236 LIVE, 75 BACKFILL)",
 "DONE","2026-07-31","LIVE = ONE rec_date into ONE expiry = 1 market outcome, no significance possible. EV rule validated: 'trap' group wins MORE (96.8%) but returns 9x LESS (+1.89% vs +16.95%)","a910cb1","P1","-"),
(10,"2026-07-31","Claude","Data","Cash-secured put capital recorded as 0",
 "risk=None for CSP in both scanners -> capital=0 -> removed from every return denominator",
 "DONE","2026-07-31","Fixed at source (risk=(K-credit)*100) in both scanners; 89 historical rows repaired. Also silently understated 'Capital at risk' on the dashboard","5c76fe7","P0","-"),
(11,"2026-07-31","Claude","Analysis","BACKFILL '+136.8% return' claim",
 "Presented as confirming the EV-rule inversion",
 "WITHDRAWN","2026-07-31","Artifact of the CSP capital=0 bug. Restated on real capital: BACKFILL +18.79% vs LIVE +14.84% - they AGREE. No inversion in BACKFILL","5c76fe7","P1","-"),
(12,"2026-07-31","User","Strategy","Exit at 50% of max profit - better?",
 "Standard premium-selling practice; reconstructed from real daily chains, ask-side closes",
 "DONE","2026-07-31","LOSES money: -$19,065 (-0.58% on capital). 73% hit target early. Worst loss IDENTICAL - zero downside protection. Caps winners, does nothing for losers","-","P2",
 "Revisit once a LOSING cohort exists; sim excludes capital-redeployment benefit"),
(13,"2026-07-31","Claude","Strategy","Cash-secured puts are the capital sink",
 "84 CSPs absorb 95.9% of capital, return 1.52%; spreads return ~14.8%",
 "DONE","2026-08-02",
 "FIXED via the concentration cap. Measured: across all 225 open recs, 75 CSPs hold 98.2% of capital at median RoR 1.10% vs 24.42% for spreads (median CSP needs $32,522 vs $164). Under 'Highest POP' ranking CSPs took 49.7% of a $10k basket; the new 'Max per position (% of capital)' cap (default 25%) drops that to 0% without banning CSPs outright - they simply cannot exceed their share","a910cb1","P1",
 "Filter or shrink CSPs in the basket sizer"),
(14,"2026-07-31","Claude","Strategy","Basket sizer deploys only 27% of $10k",
 "EV ranking picks cheap spreads, leaves $7,276 idle -> -0.21% after fees vs +21.56% at full deployment",
 "DONE","2026-08-02",
 "FIXED. Root cause was `_lots = 1` for the EV/POP rules, so the basket took one contract of each cheap spread regardless of account size. Added a scale-up pass (add lots in ranked order to a Target deployment %, skipping premium traps) plus a per-position concentration cap. Measured: $10k 45.9% -> 90.2% deployed (EV +222 -> +591); $50k 9.2% -> 90.0%; $250k 1.8% -> 90.3%, largest position always <=25%. Verified in the live DOM: $9,020 deployed, 90% of capital, zero exceptions","-","P1","Fix the deployment flaw"),
# ---------------------------------------------------------------- regime / hedging
(15,"2026-07-31","User","Strategy","Performance during market falls",
 "Claimed impossible pre-2026-07; user pushed back - correctly",
 "DONE","2026-07-31","I was WRONG. Outcomes depend only on the underlying, and stock_history reaches 1990. A 6-7% index drop = -100% on the spread. SPY 2022 = -179% on risk, erasing 2019+2021 combined","cb53358","P1","-"),
(16,"2026-07-31","User","Strategy","Index vs individual stocks - which is better?",
 "Common window 2020-08 to 2026-07 so every ticker sees 2022",
 "DONE","2026-07-31","NOT SIGNIFICANT (t=1.49, p=0.186). Withdrew a contaminated ranking - 6 tickers lacked pre-2020-07 history and skipped COVID. Within-stock dispersion dwarfs the index/stock gap","cb53358","P1","-"),
(17,"2026-07-31","User","Strategy","Hedge strategies for continuous income",
 "Tail put / collar / vol filter / wide-OTM tested on real crash cycles",
 "DONE","2026-07-31","Tail puts are -EV once REAL skew is priced (SPY 10% OTM = 1.80x ATM IV). TAIL PUT -506%, COLLAR -4016%. WIDE OTM best (+271%). Position sizing is the hedge, not options","d55400d","P1","-"),
(18,"2026-07-31","Claude","Methodology","Flat-vol pricing reversed every hedge conclusion",
 "First run priced all options at one ATM vol",
 "DONE","2026-07-31","Measured real skew in our own chains, refit, and TAIL PUT went +223% -> -506%. Caught before reporting as final","d55400d","P0","-"),
# ---------------------------------------------------------------- external tools
(19,"2026-07-30","User","Research","Review awesome-systematic-trading repo",
 "295 entries across 13 sections tabulated",
 "DONE","2026-07-30","Adopted volest (range-based vol estimators, shipped into /vrp) and the vectorbt pattern (60x faster, identical rank-IC)","e355ef7","P2","-"),
(20,"2026-07-30","User","Research","Private AI quant platforms survey",
 "GitHub API vitals, not marketing pages",
 "DONE","2026-07-30","QuantMuse flagged as a trap: 2,824 stars but 9 commits / 1 contributor. Category already absorbed by this project","c9da836","P2","-"),
(21,"2026-07-31","User","Research","QuantConnect / LEAN for backtests?",
 "Evaluated against limitations this session exposed",
 "DONE","2026-07-31","Worth a RESEARCH trial, not migration. Only thing carrying years of real option quotes w/ bid-ask (AlgoSeek + ThetaData). Will not fix statistical discipline","714cc33","P2",
 "Verify options data cost + coverage start dates"),
(22,"2026-07-31","User","Research","Quantiacs for backtests?",
 "Head-to-head vs QuantConnect, API-verified",
 "REJECTED","2026-07-31","NO options data at all. Toolbox 82 stars, ~7mo stale. Category mismatch for an equity-options book","7fe7c46","P3","-"),
(23,"2026-07-31","User","Research","Proven/backtested income-hedge strategies on GitHub",
 "Searched and verified claims",
 "DONE","2026-07-31","Headline '24.95% CAGR put selling' came from a repo with 0 commits / 0 contributors, and was internally contradictory. Real evidence base is CBOE BXM/PUT indices, not GitHub","-","P2","-"),
# ---------------------------------------------------------------- storage / schema
(24,"2026-07-31","User","Infra","How much disk does the data occupy?",
 "Measured empirically, one real day copied to temp DB",
 "DONE","2026-07-31","70.4 MB/trading-day -> ~18.2 GB/yr. 23 GB free = ~1.3 years runway","-","P1","-"),
(25,"2026-07-31","Claude","Infra","Drop options_daily as a duplicate",
 "Proposed on matching row counts",
 "WITHDRAWN","2026-07-31","WRONG - compared rows not schema. It has 29 extra columns and 20+ live read sites. Renaming would break /oi, ticker detail, expiry breakdown within minutes","-","P1","-"),
(26,"2026-07-31","Claude","Infra","Pre-July rows are irreplaceable",
 "Inverted my own archival advice",
 "DONE","2026-07-31","options_openbb starts 2026-07-02; options_daily starts 2025-12-12; US_data.db is EMPTY. Pre-July rows are the ONLY copy of Dec-Jun and cannot be regenerated","-","P0",
 "Back these up FIRST, delete last"),
(27,"2026-07-31","Claude","Infra","Drop 18 dead columns from options_daily",
 "money_*, vol_rank_*, *_info, chg_oi_* - zero references anywhere",
 "QUEUED","","Measured 16.7% smaller (18.6 -> 15.5 MB/day), ~0.24 GB/yr. Zero read sites change","-","P2",
 "Verify refs, back up schema, ALTER, VACUUM"),
(28,"2026-07-31","User","Infra","Consolidate onto options_openbb only, drop options_daily",
 "Asked whether to migrate fully",
 "QUEUED","","VIABLE - only 3 of 29 cols genuinely needed (company_name/asset_type/load_date); 18 dead, 8 OHLC only 0.4% populated post-July. Catch: pre-July has no options_openbb source, so merging breaks the 'every row has real quotes' invariant","-","P2",
 "Deferred by user - revisit when read paths need refactoring anyway"),
(29,"2026-07-31","Claude","Infra","Parquet archival as standing policy",
 "Parquet is 19x denser (3.7 vs 70 MB/day)",
 "DONE","2026-08-02",
 "Built tools/archive_parquet.py: exports one parquet per trade_date, then RE-READS it and compares row count, column set and a per-column checksum against SQLite before marking the day safe_to_prune. Fingerprint sorts each column independently because parquet round-trips can reorder rows. Script NEVER deletes; --verify-only re-checks an existing archive. Archived the 5 oldest options_openbb days (883,234 rows): 320.0 MB -> 19.7 MB, 16.2x denser, 5/5 VERIFIED on readback and again from disk. Archive lives outside the repo at Options_chain_data/archive_parquet","-","P1",
 "Export + VERIFY readback before any delete"),
(30,"2026-07-30","Claude","Data","CBOE quote drifted from chain snapshot",
 "Backfilled underlying_price from a separately-timed quote",
 "DONE","2026-07-31","AAPL 6.4% / AMZN 9.4% off on their earnings evening. Switched to put-call parity off the same snapshot - internally consistent by construction","-","P1","-"),
(31,"2026-07-30","Claude","Infra","SQLite 'database is locked' killed a 734-ticker run",
 "My ad-hoc backfill collided with the EOD capture at 11%",
 "DONE","2026-07-30","busy_timeout=30s added to every writer across all capture scripts","-","P0","-"),
# ---------------------------------------------------------------- research queue
(32,"2026-07-31","Claude","Research","Inverted-edge test",
 "scn_revert / gex / left_skew have 95% CIs entirely BELOW 50%",
 "DONE","2026-08-02","NO inverted edge. scn_revert (1162 fires) spans only 17 dates; collapsing dates flips hit rate 45.3% -> 53.9%, scn_zrev 43.3% -> 53.2%. The sub-50% CIs were fire-clustering, not signal. gamma_pin p=0.019 is 1 of 13 models tested at n=6 dates - not evidence. No inverse scanner built (closed as a save)","-","P1","-"),
(33,"2026-07-31","Claude","Research","Realistic fill modelling",
 "Every backtest except the 50%-exit test assumes mid-price fills",
 "DONE","2026-08-02","Measured on real options_openbb bid/ask for 204 settled put recs. Median short-leg spread 17.6% of mid (75th 35.8%, 90th 54.4%). Mean return mid +7.25% -> +3.81% at a midway fill = mid-pricing overstates ~2x. Win rate barely moves (94.6->94.2) - the spread eats the credit, which IS the edge. 13 recs have no positive credit at all at f=0.5","-","P1","-"),
(34,"2026-07-28","Claude","Research","Skew-panel backtest (does skew predict downside?)",
 "Gate was ~15-20 dates; rerun at 17",
 "DONE","2026-07-28","pcoi crosses significance (rank-IC -0.058, t=-3.17, correctly signed); skew25/pcvol directional but not significant. Flagged promising-not-conclusive, no scanner built",
 "42b3ab4","P2","RE-TEST: that t-stat used the pooled method later proven invalid (see ID 5)"),
(35.5,"2026-07-31","Claude","Research","Dispersion (#14)",
 "Index IV vs constituent vol",
 "QUEUED","","Separate from the skew backtest; never built, no data gate identified","-","P2","-"),
(35,"2026-07-31","Claude","Research","Formulaic alpha discovery (AlphaGen/Genetic-Alpha)",
 "Initially ranked #1 for highest ceiling",
 "REJECTED","2026-07-31","DOWNGRADED to do-not-build. It is an overfitting machine; on this sample it would find exactly the garbage the random-alpha test just found","-","P3",
 "Revisit only with walk-forward + corrected harness"),
(36,"2026-07-31","Claude","Research","Cross-strategy capital allocator",
 "Size capital across strategies by measured edge",
 "BLOCKED","","~262 graded fires/model gives a 12-point 95% CI - a 44% model is indistinguishable from 50%. Needs n~1000 (~7 months)","24c6a8b","P2","-"),
(37,"2026-07-31","User","Feature","EOD digest like InsiderFinance newsletter",
 "Reviewed each section for feasibility",
 "DONE","2026-08-02",
 "Built /digest (_eod_digest) with the 4 buildable sections: headlines, sector rotation, live option setups, unusual activity. Each states what it is worth - rotation carries the measured quadrant result, UOA is labelled descriptive-not-predictive (47.1% hit = coin flip), POP labelled model output. 5th section (AI Noteworthy Trades) omitted on purpose: needs per-print time-and-sales with aggressor side, which no free feed provides. Found and fixed 3 defects while verifying: rotation used a non-existent `item` key, 4 legacy CSP rows still carried the capital=0 bug (repaired in DB), and the RSS mix was leaking personal-finance advice columns into a market digest","-","P2","-"),
(38,"2026-07-30","User","Feature","Ollama local-LLM narrative synthesis",
 "worldmonitor pattern; free, no API cost",
 "BLOCKED","","Needs Ollama installed - your call","-","P3","-"),
(39,"2026-07-24","User","Feature","NSE / India lane",
 "Endpoints verified (bhavcopy has real delivery %)",
 "QUEUED","","Not blocked, just large: own DB, EOD script, /india command, dashboard section","-","P3","-"),
(40,"2026-07-22","User","Research","Kronos foundation model",
 "Open-source K-line foundation model",
 "BLOCKED","","Pretrained through an unpublished cutoff - backtesting on our history is contaminated and would manufacture a fake edge","-","P3","-"),
# ------------------------------------------------- added on re-audit vs all 76 commits (2026-08-01)
(41,"2026-07-31","Claude","Strategy","~100% POP recommendations were fake",
 "Some recs showed POP=100, which is impossible",
 "DONE","2026-07-31","Root-caused and fixed; fake-POP recs also excluded from the capital basket","2a172ba","P0","-"),
(42,"2026-07-31","User","Feature","Capital-deployed basket sizing (start with $10k)",
 "User: if i want to start with 10k, what should i take and track them",
 "DONE","2026-07-31","Basket sizer with persistent tracking; then found it deploys only 27% of capital (see ID 14)","40c2e31","P1","-"),
(43,"2026-07-31","Claude","Strategy","Per-row EV verdict on the basket + debit-spread fix",
 "paid% vs need% shown per row",
 "DONE","2026-07-31","Debit spreads computed max profit wrongly; fixed. This is the column later validated by the 311-rec backtest","9b80d13","P1","-"),
(44,"2026-07-30","User","Data","Stop using Yahoo for stock_daily",
 "User: why are we using yahoo, use only bb",
 "DONE","2026-07-30","Sourced close from BB own underlying_price; zero extra API calls, immune to yfinance rate-limiting","99945d1","P1","-"),
(45,"2026-07-30","Claude","Research","A18 short-interest validation",
 "Was blocked ~2-3 months on history accruing",
 "DONE","2026-07-30","UNBLOCKED ~7mo early via free Nasdaq historical API. Honest NULL result: rank-IC +0.017 (t=+0.65), no edge. Kept descriptive","49fd0c2","P2","-"),
(46,"2026-07-29","User","Feature","Rotation Tracker drilldown",
 "User: if sectors in green like health care it should point me stocks",
 "DONE","2026-07-29","Click the chart bubble directly to drill sector into constituent stocks","2f836c8","P2","-"),
(47,"2026-07-29","User","Feature","Add Victoria Fernandez / CrossMark to 13F tracker",
 "User asked to add her portfolio",
 "DONE","2026-07-29","CrossMark Global Holdings added to _EDGAR_FUNDS","0279b7a","P3","-"),
(48,"2026-07-29","User","Bug","FOMC info missing from the wrap",
 "User: where is today FED meeting info, why that missing",
 "DONE","2026-07-29","Added FOMC-today flag, full indices table, and a major-world-events scan (wars/bombings/climate per user ask)","7bdf903","P1","-"),
(49,"2026-07-28","User","Feature","M2 money supply + gold/dollar in daily alerts",
 "User asked for money-supply insight and its relation to gold/dollar",
 "DONE","2026-07-29","Added to Market Overview and the daily brief with a 5y chart. Computed the REAL correlation (weak: gold -0.10, DXY -0.03) instead of asserting the popular narrative. Later added QoQ + pipe-table format","d273612 a57abca 8e89734","P2","-"),
(50,"2026-07-28","User","Feature","Watchlist: company names + edit capability",
 "User: charts dont have names, also after ticker place the name",
 "DONE","2026-07-29","Name column, chart titles, and an Edit panel (asset class / target / note)","11352fc","P2","-"),
(51,"2026-07-28","Claude","Strategy","Exit Planner liquidity gate",
 "BS next-day model recipe: vol>=100, spread<=8%",
 "DONE","2026-07-28","Validated recipe wired into Day L-H scenario pricing; thin legs flagged","7f87e11","P2","-"),
(52,"2026-07-28","Claude","Feature","A10: consolidate recurring alerts into one message",
 "~5 recurring pushes filling the chat",
 "DONE","2026-07-28","position_monitor + intraday_alert folded into ONE edited status message per day. Trigger-only alerts left separate on purpose (they are genuine news)","1563870","P2","-"),
(53,"2026-07-28","User","Feature","README demo video of the app + Telegram",
 "User: i want video not GIF, also all pages",
 "DONE","2026-07-28","Recorded a real MP4 tour of all 35 dashboard pages via Playwright; kept GIF markdown links per user preference","70416a8 973346e 04beb6e","P3","-"),
(54,"2026-07-28","Claude","Feature","Verdict-grounding RAG",
 "Retrieve news/sentiment/13F context for a fired verdict",
 "DONE","2026-07-28","Extended corpus, deduped the MCP/bot RAG core into one implementation, wired into the UI. Verified live via Playwright","3e14b29","P2","-"),
(55,"2026-07-27","User","Bug","OI charts not appearing",
 "User: why is not getting two pictures of charts in OI, we build that",
 "DONE","2026-07-27","Two causes: Telegram flood control on rapid photo sends, then the REAL one, wired into the wrong handler","b191f43 d84e40f","P1","-"),
(56,"2026-07-27","User","Bug","Chart text illegible / spot line invisible",
 "User: color is light and cant see that, and: $61 expected day, what is this",
 "DONE","2026-07-27","Fixed an overlapping annotation collision, then the invisible white spot-line/bar on a white background; bumped text sizes","46a43fb 4e3c2a2","P2","-"),
(57,"2026-07-27","User","Bug","Earnings call transcripts stuck a quarter behind",
 "User: earning call is not latest",
 "DONE","2026-07-27","Real bug: quarter calculation used the CURRENT in-progress quarter. Also fixed CALL HIGHLIGHTS silently showing a stale quarter","be21f08 5591370","P1","-"),
(58,"2026-07-27","User","Bug","Daily alerts trimmed / showing literal HTML",
 "User: daily alerts are trimmed, many sections are not there",
 "DONE","2026-07-27","Root cause: _tg_balance did not recognise <blockquote expandable> as an opener, so every real closer was deleted as an orphan","d4c9e51","P0","-"),
(59,"2026-07-27","User","Feature","News hyperlinks everywhere",
 "User: I asked for hyperlinks to read full news",
 "DONE","2026-07-27","URLs threaded through position alerts, OI ticker view, world news; later the /wrap catalyst quote too","9129bb2 0349c62","P2","-"),
(60,"2026-07-27","Claude","Bug","yfinance silently failing on every .info/.calendar call",
 "Found during a performance investigation",
 "DONE","2026-07-27","Root cause was environment-level: antivirus SSL interception. Built a self-healing trust fix","bd11d4f","P0","-"),
(61,"2026-07-27","User","Feature","Geopolitical / supply-chain risk flag",
 "User asked for wars, bombings, climate severity in alerts",
 "DONE","2026-07-27","Country instability flags wired into /catalysts; fires only on a real match","2331abb","P2","-"),
(62,"2026-07-26","Claude","Feature","/debate Bull-vs-Bear round + prose polish",
 "TradingAgents-inspired adversarial round",
 "DONE","2026-07-27","Structured Action/Entry/Stop/Target verdict sized off real 5d realized vol; optional LLM polish with a bulletproof fallback","69e23a0 bd74c1f","P2","-"),
(63,"2026-07-27","Claude","Feature","Persistent preference memory",
 "hermes-agent pattern",
 "DONE","2026-07-27","ticker_interest table; /debate logs views, Watchlist orders by them","c456717","P3","-"),
(64,"2026-07-25","User","Bug","Closed Positions section disappeared",
 "User: i dont see closed positions section anymore",
 "DONE","2026-07-25","Real cause: a stray unconditional st.stop()","4924a35","P1","-"),
(65,"2026-07-26","Claude","Infra","Delete confirmed-dead _lib modules",
 "A16 orphan triage",
 "DONE","2026-07-26","Removed 3 modules with zero call sites after verifying the underlying tables stay alive via separate direct code","d47908d","P3","-"),
(66,"2026-07-27","Claude","Feature","A5 OI buy-to-open vs write-to-open classifier",
 "Rising OI alone is ambiguous",
 "DONE","2026-07-27","OI x price self-join. Verified live: several BULLISH call builds are actually >95% write-to-open","bc5373e","P2","-"),
(67,"2026-07-27","Claude","Infra","Audit-table closeout (A1/A11/A12/A15/A16)",
 "Several rows were stale or already fixed",
 "DONE","2026-07-27","Exit Planner reorder via st.empty(); Market Overview collapsed; Portfolio page verified live via Playwright (first time ever opened); A16/A1 confirmed false alarms","3bc98b1 7ca1296 b66d220 5c267a2 d3895e8","P2","-"),
(68,"2026-08-01","User","Process","Maintain an Excel tracker of ideas/questions",
 "One row per idea, actioned flag, date added, date actioned",
 "DONE","2026-08-01","Built IDEA_TRACKER.xlsx + regenerator; force-added past .gitignore with approval. Re-audited against all 76 session commits and added ~30 missing rows","8187857","P2","-"),
(69,"2026-07-27","User","Question","Do we have MCP connectors (GitHub/Brave/Supabase/Notion/Figma)?",
 "Asked which connectors are available",
 "DONE","2026-07-27","Only claude.ai Google Drive is configured and it needs auth. Git/GitHub MCP judged unnecessary since plain git/gh CLI covers this workflow","-","P3","-"),
(70,"2026-07-31","User","Question","Intraday alerts not arriving, only anomaly ones",
 "User asked why",
 "DONE","2026-07-31","Tabulated all 24 scheduled jobs and their trigger conditions; most are market-hours or state-change gated","-","P2","-"),
(71,"2026-07-28","Claude","Infra","Sidebar rebrand to WallStreet_Agentic_TradeOps",
 "Repo renamed from NYSE_DATA",
 "DONE","2026-07-28","Heading text + font size","8c46bb8 f414493","P3","-"),
# ------------------------- added 2026-08-01: repo asks that were done but never tracked
(72,"2026-07-24","User","Research","Scraping tools: which is best? (crawl4ai / firecrawl)",
 "User gave both repos: unclecode/crawl4ai and firecrawl/firecrawl, asked to analyse, plan, then do",
 "DONE","2026-07-24","Shortlisted Firecrawl (~155k stars, cloud or self-host) vs Crawl4AI (~75k, Playwright-based, LLM-ready markdown) vs Playwright (already installed). Recommendation: Playwright for one-off JS pages, Crawl4AI only if recurring structured scraping starts. Neither installed - not needed. Written up in docs/SCRAPING_TOOLS.md",
 "-","P3","-"),
(73,"2026-07-24","Claude","Methodology","TradingKey was assumed JS-gated - it was not",
 "The reason the scraping-tool question came up at all",
 "DONE","2026-07-24","Checked the raw HTML properly: page is server-rendered (Next.js SSR) and ships all 88 investors as a JSON blob. Plain requests.get + ?page=N works. No browser needed. Lesson recorded: verify the raw HTML before reaching for browser automation",
 "-","P2","-"),
(74,"2026-07-24","User","Research","worldmonitor / TradingAgents / hermes-agent",
 "User gave 3 repos as a possible AI hedge fund desk architecture",
 "DONE","2026-07-27","Wholesale adoption rejected (TradingAgents needs paid LLM calls per decision; the other two are large standalone products). Scoped 4 smaller pieces instead - 3 shipped (Bull-vs-Bear round, prose polish, geopolitical flag, preference memory), 1 deferred (Ollama, needs install)",
 "69e23a0 bd74c1f 2331abb c456717","P2","-"),
(75,"2026-07-25","User","Feature","virattt/ai-hedge-fund",
 "Multi-persona investor ensemble",
 "DONE","2026-07-25","Borrowed the PATTERN, not the dependency: built the AI Hedge Fund dashboard page as a free/deterministic rule-based ensemble (no LLM keys, no paid data). Confirmed in the Part 2 survey as already-absorbed",
 "-","P2","-"),
(76,"2026-08-02","Claude","Strategy","Scanner liquidity gate + realistic fill",
 "Direct follow-up to the ID 33 fill measurement",
 "DONE","2026-08-02","Added _hp_fill (shared by _hiprob_scan and _hiprob_scan_asof): rejects legs with no two-sided market or spread >35% of mid, and prices shorts at the sell side / longs at the buy side (f=0.50) instead of the mid. Verified on the 2026-07-21 universe: all 15 setups still qualify, credits fall a median 7% (up to 31% on two-leg spreads, 0.3% on index CSPs). Gate bites where measured: EFA rejects 79% of legs, SPXS 75%, BIIB 53%, liquid mega-caps 0%. Live yfinance path smoke-tested",
 "-","P1","-"),
(78,"2026-08-02","User","Feature","GEX Master Trading co-pilot (Nick Ireland framework)",
 "User pasted the full two-mode co-pilot spec; Claude initially read it as forwarded reading material and did NOT action it - caught by the user the next day",
 "DONE","2026-08-02","Built /gexplan (Mode 1 pre-market blueprint) and /gexcheck (Mode 2 live execution filter). Improves on the source prompt by auto-filling gamma flip / call wall / put wall / control node from our OWN captured chain via _compute_gex + analyze_inst_signals, instead of pasting them by hand from SpotGamma. Adds the daily 9/21/50 EMA trend gate (Bullish/Bearish/Tangled -> no-trade) and the 3-part checklist as a hard gate. Fixed a real defect during verification: the control node was picked from walls banded around the MEDIAN strike, giving SPY a $550 node against a $756 spot - now constrained to +/-10% of spot. Carries an honesty footer: on this DB the gex model hit 40.2% over 107 fires, p=0.21 de-overlapped, so levels are structure not forecast; the value is the filter",
 "-","P2","-"),
(79,"2026-08-02","Claude","Process","A pasted spec was treated as reading material, not a task",
 "The GEX prompt arrived as a forwarded WhatsApp-style thread and was answered with a critique instead of being built or logged",
 "DONE","2026-08-02","Owned and corrected. Lesson: when the user pastes a spec, log it in the tracker even if unsure whether to build it - the tracker exists precisely so an ask cannot be silently dropped. Nothing else from that day was missed (re-checked the session)",
 "-","P2","-"),
(80,"2026-08-02","User","Feature","GEX co-pilot: complete the checklist + merge the gamma pages",
 "User asked whether everything in the GEX spec existed in BOTH Telegram and Streamlit, then to merge the duplicate gamma tabs",
 "DONE","2026-08-02","Audit found 3 gaps: price-action and volume legs were only PRINTED as requirements (so a GREEN LIGHT could fire on structure alone), no 0DTE/weekly expiry choice, and NOTHING in Streamlit. All closed. (1) _gex_confirm reads real 1m bars from US_intraday.db - volume vs 20-bar mean (>=1.5x) and rejection-candle / 9-EMA-reclaim detection; tickers outside the 32-name intraday lane report UNVERIFIED and FAIL the leg rather than pass unseen. (2) expiry basis selector: SPY reads TRENDING -16M on the monthly but PINNING +43M on 0DTE - opposite regimes at the same moment. (3) merged the two dedicated pages (GEX Profile, Gamma Wall Advisor) into one 'GEX Command' page with a sidebar sub-view + new Co-Pilot view; GEX sections inside OTHER pages left untouched per user instruction. Verified in live DOM",
 "-","P2","-"),
(81,"2026-08-02","User","Research","Can anything be salvaged from GEX?",
 "Asked after being told the GEX levels have no measured edge",
 "DONE","2026-08-02","Found we had been testing the WRONG CLAIM: every prior test scored GEX on DIRECTION (40.2% hit, p=0.21), but dealer-hedging theory makes no directional claim - it predicts realised VOL. Tested properly. First harness gave IC +0.146 t=+3.85 p=0.0008 but FAILED its sanity gate (15% of random signals passed) - discarded, not reported. Causes: cross-ticker level effect + vol persistence. Framed as the theory actually states it (time-series, per-underlying): SPY high-GEX days -> 10.2% fwd vol vs 14.0% after low-GEX; QQQ 16.6% vs 23.6%. Correctly signed and economically large (~27-30% lower) but n=12/bucket, p=0.06-0.09, NOT significant. Conclusion: keep GEX as a volatility-REGIME read for sizing/strategy choice, never for direction; do not build a directional GEX scanner. Re-test at >=40 obs per bucket. ADOPTED.md Part 10",
 "-","P2","Re-test when ~8 more months accrue, or sooner using options_openbb real gamma"),
]

def build():
    wb = Workbook(); ws = wb.active; ws.title = "Ideas & Questions"
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    status_fill = {
        "DONE":      PatternFill("solid", fgColor="C6EFCE"),
        "QUEUED":    PatternFill("solid", fgColor="FFF2CC"),
        "BLOCKED":   PatternFill("solid", fgColor="F8CBAD"),
        "REJECTED":  PatternFill("solid", fgColor="D9D9D9"),
        "WITHDRAWN": PatternFill("solid", fgColor="FFC7CE"),
        "USER-ACTION": PatternFill("solid", fgColor="BDD7EE"),
    }
    prio_font = {"P0": Font(bold=True, color="9C0006"), "P1": Font(bold=True, color="974706")}

    ws.append(COLS)
    for i, cell in enumerate(ws[1], 1):
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in ROWS:
        ws.append(list(r))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(COLS)):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
        st = row[COLS.index("Status")].value
        if st in status_fill:
            row[COLS.index("Status")].fill = status_fill[st]
            row[COLS.index("Status")].font = Font(bold=True)
        pr = row[COLS.index("Priority")].value
        if pr in prio_font:
            row[COLS.index("Priority")].font = prio_font[pr]

    widths = {"ID":5,"Date Added":12,"Raised By":10,"Category":13,"Idea / Question":34,
              "Detail":42,"Status":12,"Date Actioned":13,"Outcome / Evidence":62,
              "Commit":10,"Priority":8,"Next Step":34}
    for i, c in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 18)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"

    # summary sheet
    s = wb.create_sheet("Summary")
    s.append(["Status", "Count"]);
    counts = {}
    for r in ROWS: counts[r[6]] = counts.get(r[6], 0) + 1
    for k, v in sorted(counts.items(), key=lambda x: -x[1]): s.append([k, v])
    s.append([]); s.append(["Category", "Count"])
    cc = {}
    for r in ROWS: cc[r[3]] = cc.get(r[3], 0) + 1
    for k, v in sorted(cc.items(), key=lambda x: -x[1]): s.append([k, v])
    for row in s.iter_rows():
        for c in row:
            if c.row == 1 or (c.value in ("Category",) and c.column == 1):
                c.font = Font(bold=True)
    s.column_dimensions["A"].width = 16; s.column_dimensions["B"].width = 8

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}  ({len(ROWS)} rows)")
    for k, v in sorted(counts.items(), key=lambda x: -x[1]):
        print(f"   {k:12s} {v}")

if __name__ == "__main__":
    build()
